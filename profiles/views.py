from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.views.generic.base import TemplateView
from django.views.generic import ListView, DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth import login
from django.contrib.auth.mixins import UserPassesTestMixin
from django.db.models import Q
# Messages
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
# Display message when user log in/out
from django.contrib.auth.signals import user_logged_out, user_logged_in
from django.dispatch import receiver
# Translation
from django.utils.translation import gettext_lazy as _
from django.utils.translation import get_language
# GeoDjango
from django.contrib.gis.db.models.functions import Distance
from django.contrib.gis.geos import fromstr
from django.contrib.gis.geos import GEOSGeometry
# Sending Emails
from django.core.mail import EmailMessage, BadHeaderError, EmailMultiAlternatives
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
# Sites
from django.contrib.sites.shortcuts import get_current_site
# Exporting Excel files
from datetime import datetime
from datetime import date
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
#import string
from openpyxl.drawing.image import Image
from openpyxl.chart import PieChart, Reference
from openpyxl.worksheet.pagebreak import RowBreak, Break, ColBreak
import os
from itertools import groupby
from operator import attrgetter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
# Calendar
from core.utils.calendar import CalEvents, get_date, prev_month, next_month, get_cal_locale
from django.utils.safestring import mark_safe
# AJAX request
from django.views.decorators.csrf import csrf_exempt

from profiles.models import Profile, CustomUser
from activities.models import Activity, Grade
from outings.models import Outing
from availabilities.models import Availability
from profiles.forms import ProfileCreationForm, AccountForm, ProfileForm, ContactProfileForm

# from django.contrib.gis.geos import Point
# longitude = 8.191788
# latitude = 48.761681
# user_location = Point(longitude, latitude, srid=4326)


# VIEWS FOR ANONYM WEB SURFERS

class ProfileListView(ListView):
    '''
    Display the all the public profiles by default.
    If a filter is activated it displays the profiles according to the filter.
    '''
    template_name = "profiles/profile_list.html"
    nb_of_results = 0

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['all_activities'] = Activity.objects.all()
        context['nb_of_results'] = self.nb_of_results
        context['selected_activities'] = self.request.GET.getlist('a', None)
        context['around_me'] = self.request.GET.get('around_me', None)
        context['start_date'] = self.request.GET.get('start_date', None)
        context['end_date'] = self.request.GET.get('end_date', None)
        context['availability_area_geo'] = self.request.GET.get('availability_area_geo', None)
        return context

    def get_queryset(self):
        start_date = self.request.GET.get('start_date', None)
        if start_date:
            from_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = self.request.GET.get('end_date', None)
        if end_date:
            till_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        selected_activities = self.request.GET.getlist('a', None)
        activities_in_current_language = 'activities__name_{}'.format(get_language())
        around_me = self.request.GET.get('around_me')
        availability_area_geo = self.request.GET.get('availability_area_geo',None)
        if availability_area_geo:
            availability_area_geo_obj = GEOSGeometry(availability_area_geo)
        user_authenticated = self.request.user.is_authenticated

        q = Profile.objects.select_related('user').prefetch_related('activities').filter(public_profile='True')
        
        if user_authenticated:
            q = q.exclude(user=self.request.user)

        if user_authenticated and self.request.user.is_profile:
            if self.request.user.profile.location:
                user_loc = self.request.user.profile.location
                q = q.annotate(distance=Distance('location', user_loc)).order_by('distance', '-last_update')
                if around_me:
                    q = q.filter(location__distance_lte=(user_loc, 50000))

        if start_date and end_date:
            q = q.filter(user__availability__start_date__lte=till_date, user__availability__end_date__gte=from_date)
        elif start_date:
            q = q.filter(user__availability__end_date__gte=from_date)
        elif end_date:
            q = q.filter(user__availability__start_date__lte=till_date, user__availability__end_date__gte=date.today())

        if selected_activities:
            q = q.filter(
                eval(' | '.join(f'Q({ activities_in_current_language }="{ selected_activity }")' for selected_activity in selected_activities)),
                public_profile='True'
            )
        if availability_area_geo:
            q = q.filter(availability_area_geo__intersects=availability_area_geo_obj)

        if (start_date or end_date or selected_activities) and user_authenticated and self.request.user.is_profile:
            if self.request.user.profile.location:
                q = q.distinct('distance', 'last_update', 'user_id')
            else:
                q = q.distinct('last_update', 'user_id')
        elif start_date or end_date or selected_activities:
            q = q.distinct('last_update', 'user_id')

        self.nb_of_results = len(q)

        return q


class ProfileDetailView(DetailView):
    '''A view for the web surfer to detail a specific profile'''

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        profile = get_object_or_404(Profile.objects.prefetch_related('user__outing_set__activities', 'user__availability_set', 'grades__activity'), user__username=self.kwargs['username'])
        context['profile'] = profile
        if profile.availability_area_geo is not None:
            poly_tuple = profile.availability_area_geo.coords[0]
            context['availability_area_geo_poly'] = [[i[0], i[1]] for i in poly_tuple] or None
        # Calendar Context
        if profile.user.outing_set.all() or profile.user.availability_set.all():
            d = get_date(self.request.GET.get('month', None))
            cal = CalEvents(year=d.year, month=d.month, profile=self.kwargs['username'], locale=get_cal_locale(get_language()))
            html_cal = cal.formatmonth(withyear=True)
            context['cal_events'] = mark_safe(html_cal)
            context['prev_month'] = prev_month(d)
            context['next_month'] = next_month(d)
        return context

    def get_object(self):
        return get_object_or_404(Profile, user__username=self.kwargs['username'])


def ContactProfileView(request, **kwargs):
    '''A view for the web surfer to send an email to the user through a contact form.'''
    user_contacted_username = kwargs['username']
    user_contacted = CustomUser.objects.get(username=user_contacted_username)
    if user_contacted.first_name and user_contacted.last_name:
        user_contacted_name = user_contacted.first_name + ' ' + user_contacted.last_name
    elif user_contacted.first_name:
        user_contacted_name = user_contacted.first_name
    elif user_contacted.last_name:
        user_contacted_name = user_contacted.last_name
    else:
        user_contacted_name = user_contacted.username
    user_contacted_phone_number = CustomUser.objects.get(username=user_contacted_username).phone_number

    if request.method == 'POST':
        form = ContactProfileForm(request.POST)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            subject_prefixed = '[bourseauxcompagnons] ' + subject 
            from_email = form.cleaned_data['from_email']
            recipients = [user_contacted.email]
            contact_message = form.cleaned_data['message']
            current_site = get_current_site(request)
            html_message = render_to_string(
                'profiles/contact_profile_email_inline.html',
                {'profile_contacted': user_contacted_name,
                'profile_making_contact': request.user.username,
                'message': contact_message,
                'site_name': current_site.name,
                'domain': current_site.domain,
                'protocol': "https" if request.is_secure() else "http",
                }
            )
            # plain_message = strip_tags(html_message) # I finally prefer to create a contact_profile_email_plain.html than just strip the tags off that result in a shitty email.
            plain_message = render_to_string(
                'profiles/contact_profile_email_plain.html',
                {'profile_contacted': user_contacted_name,
                'profile_making_contact': request.user.username,
                'message': contact_message,
                'site_name': current_site.name,
                'domain': current_site.domain,
                'protocol': "https" if request.is_secure() else "http",
                }
            )
            try:
                email = EmailMultiAlternatives(
                    subject_prefixed,
                    plain_message,
                    "Bourse aux compagnons <contact@bourseauxcompagnons.fr>",
                    recipients,
                    bcc=['contact@bourseauxcompagnons.fr'],
                    reply_to=[from_email],
                )
                email.attach_alternative(html_message, "text/html")
                email.send()
            except BadHeaderError:
                return HttpResponse('Invalid header found.')
            msg = _('Your message has been sent by email to {}. Hopefully, he or she will get back to you soon! :) <a href="{}">Go back to the profiles page</a>').format(user_contacted_name, reverse_lazy('profiles:list'))
            messages.add_message(request, messages.SUCCESS, msg)
            return redirect('profiles:detail', username=user_contacted_username)
    else:
        prepopulated_fields = {'subject': _('Initial contact'),
                                'message': _('Hi ')+user_contacted_name+',\n'}
        if request.user.is_authenticated:
            prepopulated_fields.update({'from_email': request.user.email})
        form = ContactProfileForm(initial=prepopulated_fields)

    return render(request, "profiles/contact_profile_form.html", {'form': form, 'profile_contacted': user_contacted_name, 'profile_phone_number': user_contacted_phone_number})


class ProfileRegisterView(SuccessMessageMixin, CreateView):
    '''A view for the web surfer to create a profile'''
    form_class = ProfileCreationForm
    template_name = 'profiles/profile_register.html'
    success_url = reverse_lazy('my-profile')
    success_message = _("WARNING: your profile is public by default. That is to say that it will be displayed "
                        "in the profile list available on bourseauxcompagnons. Click on "
                        "'Update my profile' to complete it or make it private."
                        )
    
    def form_valid(self, form):
        '''
        1. Save the valid form (useless because already saved through ProfileCreationForm but it provides me with self.object which is the customuser)
        2. Send an confirmation email to the new profile
        3. Login the new profile
        4. Redirect to success_url.

        No authentication.
        '''
        response = super().form_valid(form)

        subject=_("Profile registered")
        subject_prefixed = _("[Account] {}").format(subject)
        recipients = [self.object.email]
        html_message = render_to_string('profiles/profile_register_email_inline.html', {'customuser': self.object})
        # plain_message = strip_tags(html_message) # I finally prefer to create a profile_register_email_plain.html than just strip the tags off that result in a shitty email.
        plain_message = render_to_string('profiles/profile_register_email_plain.html', {'customuser': self.object})
        send_mail(subject_prefixed, plain_message, "Bourse aux compagnons <contact@bourseauxcompagnons.fr>", recipients, html_message=html_message)
        
        login(self.request, self.object)
        
        return response


# VIEWS FOR AUTHENTICATED PROFILES

class ProfileHomepageView(UserPassesTestMixin, TemplateView):
    template_name = 'profiles/my_profile.html'

    def test_func(self):
        return self.request.user.is_authenticated

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        outings = Outing.objects.filter(author=self.request.user)
        availabilities = Availability.objects.filter(author=self.request.user).exists()
        context['outings'] = outings
        context['availabilities'] = availabilities
        # Calendar Context
        d = get_date(self.request.GET.get('month', None))
        cal = CalEvents(year=d.year, month=d.month, profile=self.request.user.username, locale=get_cal_locale(get_language()))
        html_cal = cal.formatmonth(withyear=True)
        context['cal_events'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d)
        context['next_month'] = next_month(d)
        return context


class ProfileUpdateView(UserPassesTestMixin, UpdateView):
    '''A view for the authenticated profile to update its profile.'''
    form_class = ProfileForm

    def test_func(self):
        return self.request.user.username == self.kwargs['username']

    def get_object(self):
        return get_object_or_404(Profile, user__username=self.kwargs['username'])
    
    # def form_valid(self, form):
        # coordinates = form.cleaned_data['location'].split(',')
        # print(coordinates)
        # form.instance.location = Point(float(coordinates[0]),float(coordinates[1]))
        # return super().form_valid(form)


# @csrf_exempt # I used the csrf exemption before april 2022 when I added the CSRF protection on the ajax request
def load_grades(request):
    checked_activities = request.POST.getlist('a')
    checked_grades = request.POST.getlist('g')
    if checked_activities:
        grades = Grade.objects.filter(eval(' | '.join(f'Q(activity="{ activity }")' for activity in checked_activities)))
    else:
        grades = Grade.objects.none()
    checked_grades = list(map(int, checked_grades))
    return render(request, 'profiles/grades_dropdown_list_options.html', {'grades': grades, 'checked_grades': checked_grades})


class AccountUpdateView(UserPassesTestMixin, UpdateView):
    '''A view for the authenticated profile to update its profile account settings.'''
    form_class = AccountForm
    template_name = 'registration/account_form.html'
    success_url = reverse_lazy('my-profile')

    def test_func(self):
        return self.request.user.username == self.kwargs['username']

    def get_object(self):
        return get_object_or_404(CustomUser, username=self.kwargs['username'])


class AccountDeleteView(UserPassesTestMixin, SuccessMessageMixin, DeleteView):
    '''A view for the authenticated profile to delete its own profile and account'''
    model = CustomUser
    template_name = 'registration/account_confirm_delete.html'
    success_url = reverse_lazy('homepage')
    success_message = _("Your account and all its data are deleted. Have a good time in the outdoors!")

    def test_func(self):
        return self.request.user.username == self.kwargs['username']

    def get_object(self):
        return get_object_or_404(CustomUser, username=self.kwargs['username'])


    def form_valid(self, form):
        response = super().form_valid(form)
        subject=_("Profile deleted")
        subject_prefixed = _("[Account] {}").format(subject)
        recipients = [self.object.email]
        html_message = render_to_string('profiles/profile_deletion_email_inline.html', {'customuser': self.object})
        # plain_message = strip_tags(html_message) # I finally prefer to create a profile_deletion_email_plain.html than just strip the tags off that result in a shitty email.
        plain_message = render_to_string('profiles/profile_deletion_email_plain.html', {'customuser': self.object})
        send_mail(subject_prefixed, plain_message, "Bourse aux compagnons <contact@bourseauxcompagnons.fr>", recipients, html_message=html_message)
        return response

    def get_success_message(self, cleaned_data):
        return self.success_message.format(reverse_lazy('my-profile'))


@receiver(user_logged_in)
def when_user_logged_in(sender, request, user, **kwargs):
    if user.first_name:
        msg = _('Hi {}! You are logged in :)').format(user.first_name)
    elif user.username:
        msg = _('Hi {}! You are logged in :)').format(user.username)
    else:
        msg = _('You are logged in :)')
    messages.add_message(request, messages.SUCCESS, msg)


@receiver(user_logged_out)
def when_user_logged_out(sender, request, user, **kwargs):
    if user.first_name:
        msg = _('You have securely logged out. Thank you for visiting {}.').format(user.first_name)
    elif user.username:
        msg = _('You have securely logged out. Thank you for visiting {}.').format(user.username)
    else:
        msg = _('You have securely logged out. Thank you for visiting.')
    messages.add_message(request, messages.SUCCESS, msg)

# VIEWS FOR EXPORTING

def export_profiles_to_xlsx(request):
    '''A view to export all the profiles to an xlsx format'''
    profiles_queryset = Profile.objects.all()
    activities = Activity.objects.all()

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={date}_bac_profiles.xlsx'.format(date=datetime.now().strftime('%Y%m%d'))
    workbook = Workbook()
    
    ### STYLES ###
    # Colors
    dark_red = 'FFC00000'
    dark_blue = 'FF002060'
    # Bold
    bold_st = NamedStyle(name="bold_st")
    bold_st.font = Font(name='Calibri', bold=True, color='FF000000')
    # Main title row
    main_title = NamedStyle(name="main_title")
    main_title.font = Font(name='Calibri', bold=True, color='FF000000', size='20')
    main_title.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    main_title.border = Border(
        left=Side(border_style='thin', color='FF00008F'),
        right=Side(border_style='thin', color='FF00008F'),
        top=Side(border_style='thin', color='FF00008F'),
        bottom=Side(border_style='thin', color='FF00008F')
    )
    main_title.fill = PatternFill(fill_type='solid', fgColor='FF4080C6')
    # First row
    first_row = NamedStyle(name="first_row")
    first_row.font = Font(name='Calibri', bold=True, color='FFFFFFFF')
    first_row.alignment = Alignment(wrap_text=True, vertical='center')
    first_row.border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000')
    )
    first_row.fill = PatternFill(fill_type='solid', fgColor='FF6F6F6F')
    # Values row
    values_st = NamedStyle(name="values_st")
    values_st.font = Font(name='Calibri', color='FF000000')
    values_st.alignment = Alignment(wrap_text=True, vertical='top')
    values_st.border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000')
    )
    # Category title row
    cat_title_st = NamedStyle(name="cat_title_st")
    cat_title_st.font = Font(name='Calibri', color=dark_blue, size='16')
    cat_title_st.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
    cat_title_st.border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000')
    )
    cat_title_st.fill = PatternFill(fill_type='solid', fgColor='FFA8D18E')
    # Date format
    date_format = NamedStyle(name="date_format")
    date_format.font = Font(name='Calibri', color='FF000000')
    date_format.alignment = Alignment(wrap_text=True, vertical='top')
    date_format.border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000')
    )
    date_format.number_format='YYYY-MM-DD'
    
    ### PROFILE WORKSHEET ###
    # Creating the worksheet
    profiles_worksheet = workbook.active
    # Page setup (print settings)
    profiles_worksheet.page_setup.orientation = 'landscape' #profiles_worksheet.ORIENTATION_PORTRAIT
    profiles_worksheet.page_setup.paperSize = profiles_worksheet.PAPERSIZE_A4
    profiles_worksheet.page_setup.fitToPage = True
    profiles_worksheet.page_setup.fitToHeight = False
    profiles_worksheet.page_margins.top = 0.75 # inches (1,91 cm)
    profiles_worksheet.page_margins.bottom = 0.75 # inches (1,91 cm)
    profiles_worksheet.page_margins.left = 0.7 # inches (1,78 cm)
    profiles_worksheet.page_margins.right = 0.7 # inches (1,78 cm)
    profiles_worksheet.page_margins.header = 0.3 # inches (0,76 cm)
    profiles_worksheet.page_margins.footer = 0.3 # inches (0,76 cm)
    profiles_worksheet.print_options.horizontalCentered = True
    profiles_worksheet.sheet_view.view = 'pageBreakPreview' # or 'normal' or 'pageLayout'
    profiles_worksheet.sheet_view.zoomScale = 75
    profiles_worksheet.sheet_view.zoomScaleNormal = 75
    profiles_worksheet.sheet_view.zoomScaleSheetLayoutView = 75
    profiles_worksheet.sheet_view.zoomScalePageLayoutView = 75
    #profiles_worksheet.print_area = profiles_worksheet.dimensions
    # Title
    profiles_worksheet.title = 'Profiles'
    # First row of the table
    attributes = [
        'username',
        'first name',
        'last name',
        'public profile',
        'introduction',
        'list of courses',
        'activities',
        'birthdate',
        'last_update',
        'updated more than 1 month ago',
        'action'
    ]

    # HEADER & FOOTER
    profiles_worksheet.oddFooter.center.text = "bourseauxcompagnons.fr"
    profiles_worksheet.oddFooter.center.size = 11
    
    # Initializing variables
    row_breaks_list = []
    row = 1

    # COVER PAGE
        # Blank row
    profiles_worksheet.row_dimensions[row].height = 20*3
        # Logo
    row += 1
    profiles_worksheet.row_dimensions[row].height = 90
    dirname = os.path.dirname(os.path.dirname(__file__))
    filename = os.path.join(dirname, 'collected_static/img/icon_group_map.png')
    img = Image(filename)
    anchor = 'D2'
    profiles_worksheet.add_image(img, anchor)
        # Confidential
    CP_confidential = profiles_worksheet.cell(column=len(attributes)-1, row=row)
    CP_confidential.value = "NON CONFIDENTIEL"
    CP_confidential.font = Font(name='Calibri', bold=True, color=dark_red)
    CP_confidential.alignment = Alignment(vertical='top', horizontal='right')
        # Title
    row += 1
    CP_title = profiles_worksheet.cell(column=4, row=row)
    CP_title.value = "LISTE COMPLETE DES PROFILS\n(sans les coordonnées personnelles)"
    profiles_worksheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=len(attributes)-2)
    profiles_worksheet.row_dimensions[row].height = 409 # 409 pt is the maximum row height in Microsoft Excel
    CP_title.font = Font(name='Calibri', bold=True, color='FF000000', size='20')
    CP_title.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        # Date
    row += 1
    CP_date = profiles_worksheet.cell(column=len(attributes)-1, row=row)
    CP_date.value = "Le {}".format(date.today())
    CP_date.alignment = Alignment(vertical='top', horizontal='right')
        # Blank row
    row += 1
    profiles_worksheet.row_dimensions[row].height = 20
        # Page break
    row_breaks_list.append(Break(id=row))
    
    # OVERVIEW
        # Title
    row += 1
    O_title = profiles_worksheet.cell(column=1, row=row)
    O_title.value = "SYNTHESE"
    O_title.style = main_title
    profiles_worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(attributes))
    profiles_worksheet.row_dimensions[row].height = 40
        # Blank row
    row += 1
        # Content
    row += 1
    O_publication_date = profiles_worksheet.cell(column=1, row=row)
    O_publication_date.value = "Date d'export"
    O_publication_date.style = bold_st
    profiles_worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    O_publication_date_value = profiles_worksheet.cell(column=3, row=row)
    O_publication_date_value.value = date.today()
    row += 1
    O_nb_of_profiles = profiles_worksheet.cell(column=1, row=row)
    O_nb_of_profiles.value = "Nombre de profils"
    O_nb_of_profiles.style = bold_st
    profiles_worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    O_nb_of_profiles_value = profiles_worksheet.cell(column=3, row=row)
    O_nb_of_profiles_value.value = len(profiles_queryset)
        # Blank row
    row += 1
        # Page break
    row_breaks_list.append(Break(id=row))
    
    # TABLE
    row += 1
    first_row_of_table = row
        # Creating the first row
    for col_num, col_title in enumerate(attributes,1):
        cell = profiles_worksheet.cell(row=first_row_of_table, column=col_num, value=col_title.upper())
        cell.style = first_row
        column_dimensions = profiles_worksheet.column_dimensions[get_column_letter(col_num)]
        if col_title == 'introduction':
            column_dimensions.width = 45
        elif col_title == 'list of courses':
            column_dimensions.width = 45
        elif col_title == 'activities':
            column_dimensions.width = 35
        elif col_title == 'updated more than 1 month ago':
            column_dimensions.width = 20
        else:
            max_length = 0
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            adjusted_width = (max_length + 2)*1.2
            column_dimensions.width = adjusted_width
    
        # Freezing the first row
    # freezing_cell = "A{}".format(first_row_of_table+1)
    # profiles_worksheet.freeze_panes = profiles_worksheet[freezing_cell]
    
        # Adding the values
    values_row = first_row_of_table
    for profile in profiles_queryset:
        values_row += 1
        values = [
            (profile.user.username, values_st),
            (profile.user.first_name, values_st),
            (profile.user.last_name, values_st),
            (profile.public_profile, values_st),
            (profile.introduction, values_st),
            (profile.list_of_courses, values_st),
            (', '.join([str(i) for i in profile.activities.all()]), values_st),
            (profile.birthdate, date_format),
            (profile.last_update, date_format),
            (profile.updated_more_than_1_month_ago, values_st),
            ('A contacter', values_st)
        ]
        for col_num, (col_value, col_format) in enumerate(values, 1):
            cell = profiles_worksheet.cell(row=values_row, column=col_num, value=col_value)
            cell.style = col_format
    
        # Adding filtering
    profiles_worksheet_dim = profiles_worksheet.dimensions
    profiles_worksheet_dim_list = list(profiles_worksheet_dim)
    profiles_worksheet_dim_list[1] = str(first_row_of_table)
    profiles_worksheet_dim = "".join(profiles_worksheet_dim_list)
    profiles_worksheet.auto_filter.ref = profiles_worksheet_dim

    #     # Adding the values by category (is the profile active for 1 month ago?)
    # q_list = [profile for profile in profiles_queryset]
    # q_list = sorted(q_list, key=attrgetter('updated_more_than_1_month_ago'), reverse=True)
    # cat_dict = {"False": "Updated less than 1 month ago", "True": "Updated more than 1 month ago"}
    # for key, group in groupby(q_list, attrgetter('updated_more_than_1_month_ago')):
    #     values_row += 1
    #     print(key)
    #     O_title = profiles_worksheet.cell(row=values_row, column=1, value=cat_dict[str(key)])
    #     profiles_worksheet.merge_cells(start_row=values_row, start_column=1, end_row=values_row, end_column=len(attributes))
    #     profiles_worksheet.row_dimensions[values_row].height = 30
    #     O_title.style = cat_title_st
    #     for profile in group:
    #         values_row += 1
    #         values = [
    #             (profile.user.username, values_st),
    #             (profile.user.first_name, values_st),
    #             (profile.user.last_name, values_st),
    #             (profile.public_profile, values_st),
    #             (profile.introduction, values_st),
    #             (profile.list_of_courses, values_st),
    #             (', '.join([str(i) for i in profile.activities.all()]), values_st),
    #             (profile.birthdate, date_format),
    #             (profile.last_update, date_format),
    #             (profile.updated_more_than_1_month_ago, values_st),
    #             ('A contacter', values_st)
    #         ]
    #         for col_num, (col_value, col_format) in enumerate(values, 1):
    #             cell = profiles_worksheet.cell(row=values_row, column=col_num, value=col_value)
    #             cell.style = col_format

    # Data validation (ex: dropdow list)
    action_range = '{col}{first_row}:{col}{last_row}'.format(
        col=get_column_letter(attributes.index('action')+1),
        first_row = first_row_of_table+1,
        last_row = values_row
    )
    actions_dic = {'A contacter':'FF00FF00', 'A relancer':'FFFFFF00', 'Sélectionné':'FFFF00FF', 'Ecarté':'FF00FFFF'}
    actions_formula1 = '"' + ','.join(actions_dic.keys()) + '"'
    data_val = DataValidation(type="list",formula1=actions_formula1, allow_blank=True)
    profiles_worksheet.add_data_validation(data_val)
    data_val.add(action_range)

    # Conditional formatting
        # First name = "Yann"
    first_name_range = '{col}{first_row}:{col}{last_row}'.format(
        col=get_column_letter(attributes.index('first name')+1),
        first_row = first_row_of_table+1,
        last_row = values_row
    )
    profiles_worksheet.conditional_formatting.add(
        first_name_range,
        CellIsRule(operator='equal', formula=['"Yann"'], fill=PatternFill(bgColor="FFC7CE"))
    )

        # Actions
    for action, color in actions_dic.items():
        action_formula = '"' + action + '"'
        profiles_worksheet.conditional_formatting.add(
            action_range,
            CellIsRule(operator='equal', formula=[action_formula], fill=PatternFill(bgColor=color))
        )

    # Adding the declared page breaks to the worksheet
    profiles_worksheet.page_breaks = (RowBreak(brk=row_breaks_list), ColBreak())

    ### STAT WORKSHEET ###
    # Creating the worksheet
    stat_worksheet = workbook.create_sheet("Stat")
    
    # Distribution of profiles by activity
    activities_distribution = [('Activity', 'Number of profiles')]
    activities_in_current_language = 'activities__name_{}'.format(get_language())
    
    for activity in activities:
        activity_filter = {}
        activity_filter[activities_in_current_language] = activity
        number_of_profiles = Profile.objects.filter(**activity_filter).count()
        activities_distribution.append((str(activity), number_of_profiles))
    for row in activities_distribution:
        stat_worksheet.append(row)
    
    piechart_profiles_by_activity = PieChart()
    labels = Reference(stat_worksheet, min_col = 1, min_row = 2, max_row = activities.count() + 1)
    data = Reference(stat_worksheet, min_col = 2, min_row = 1, max_row = activities.count() + 1)
    piechart_profiles_by_activity.add_data(data, titles_from_data=True)
    piechart_profiles_by_activity.set_categories(labels)
    piechart_profiles_by_activity.title = "Profiles distribution by activities"
    piechart_profiles_by_activity.width = 15 # cm
    piechart_profiles_by_activity.height = 8 # cm
    piechart_profiles_by_activity.legend.position = 'l'
    
    piechart_profiles_by_activity.dataLabels = DataLabelList()
    piechart_profiles_by_activity.dataLabels.showPercent = True
    piechart_profiles_by_activity.dataLabels.showVal = False
    piechart_profiles_by_activity.dataLabels.showCatName = False
    piechart_profiles_by_activity.dataLabels.showLegendKey = False

    stat_worksheet.add_chart(piechart_profiles_by_activity, "D1")
    
    ### SAVING WORKBOOK ###
    workbook.save(response)

    return response